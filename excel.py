import openpyxl as xl
from openpyxl.chart import BarChart3D, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 2)       #base value row
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row,3)  #to save changed value
        corrected_price_cell.value = corrected_price

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet .cell(row,4)
        corrected_price_cell.value = corrected_price

# to make 3D chart
        """        values = Reference(sheet, min_row=2,
                            max_row=sheet.max_row,
                            min_col=2,
                            max_col=4)
                chart = BarChart3D()
                chart.add_data(values)
                chart.title = "3D Bar Chart"
                sheet.add_chart(chart, 'h2')
        """
    wb.save(filename)

print(process_workbook("storen.xlsx"))